# -*- coding: utf-8 -*-
# @time:2019/5/14 14:22
# Author:殇殇
# @file:do_excel.py
# @fuction: 封装操作Excel的读与写
from openpyxl import load_workbook

class case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.check_sql = None


class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    # 读取Excel里面的测试数据
    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        cases=[]
        for i in range(2,sheet.max_row+1):
            row_case = case () # 实例化
            row_case.case_id = sheet.cell(row=i,column =1).value
            row_case.title = sheet.cell(row=i,column=2).value
            row_case.url = sheet.cell(row=i,column=3).value
            row_case.data = sheet.cell(row=i,column=4).value
            row_case.method = sheet.cell(row=i,column=5).value
            row_case.expected = sheet.cell(row=i,column=6).value
            row_case.check_sql = sheet.cell(row=i,column=9).value
            cases.append(row_case)
        wb.close()
        return cases

    # 将测试结果写回到Excel中
    def write_back(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        wb.save(self.file_name)
        wb.close()

